import warnings
import os
from datetime import datetime, date

import openpyxl


# Índices padrão das colunas (0-based). Usados como fallback se o cabeçalho
# não contiver as colunas de parcela de forma identificável.
COL_DATA_PAGAMENTO        = 0
COL_DATA_LANCAMENTO       = 1
COL_ESTABELECIMENTO       = 2
COL_TIPO_LANCAMENTO       = 3
COL_FORMA_PAGAMENTO       = 4
COL_BANDEIRA              = 5
COL_VALOR_BRUTO           = 6
COL_TAXA_TARIFA           = 7
COL_VALOR_LIQUIDO         = 8
COL_STATUS_PAGAMENTO      = 9
COL_DATA_VENDA            = 11
COL_HORA_VENDA            = 12
COL_AUTORIZACAO           = 15
COL_NSU_DOC               = 16
COL_COD_VENDA             = 17
COL_NUM_PARCELA           = 27
COL_TOTAL_PARCELAS        = 28
COL_TAXA_TOTAL            = 36
COL_VALOR_TAXA_MDR        = 39
COL_VALOR_TOTAL_TRANSACAO = 57

HEADER_MARKER = 'Data de pagamento'


def _parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_decimal(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(',', '.'))
    except (ValueError, TypeError):
        return None


def _parse_int(value):
    """Converte para int; aceita float (1.0), string int ('1') e string float ('1.0')."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    try:
        s = str(value).strip()
        if not s:
            return None
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _str(value, maxlen=None):
    if value is None:
        return None
    s = str(value).strip()
    if maxlen:
        s = s[:maxlen]
    return s or None


def _find_header_row(ws):
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True)):
        if row and row[0] and str(row[0]).strip() == HEADER_MARKER:
            return i  # 0-based index
    return None


def _auto_col(header_vals, *must_contain, exclude=()):
    """
    Retorna o índice (0-based) da primeira coluna cujo cabeçalho contém
    todos os termos de 'must_contain' e nenhum de 'exclude' (case-insensitive).
    Retorna None se não encontrar.
    """
    must = [m.lower() for m in must_contain]
    excl = [e.lower() for e in exclude]
    for i, v in enumerate(header_vals):
        if v is None:
            continue
        h = str(v).strip().lower()
        if all(m in h for m in must) and not any(e in h for e in excl):
            return i
    return None


def _detect_parcel_cols(header_vals):
    """
    Tenta detectar automaticamente as colunas de parcela no cabeçalho.
    Retorna (col_num_parcela, col_total_parcelas) usando fallback nos índices
    padrão se não encontrar pelos nomes.
    """
    col_num = (
        _auto_col(header_vals, 'parcela', exclude=('total', 'quantidade', 'qtd'))
        or COL_NUM_PARCELA
    )
    col_tot = (
        _auto_col(header_vals, 'total', 'parcela')
        or _auto_col(header_vals, 'quantidade', 'parcela')
        or _auto_col(header_vals, 'qtd', 'parcela')
        or COL_TOTAL_PARCELAS
    )
    return col_num, col_tot


def import_excel(file_path):
    """
    Lê a planilha Cielo e retorna (id_arquivo, lista_de_dicts).
    Cada dict corresponde a uma linha de dados pronta para INSERT.
    """
    warnings.filterwarnings('ignore')

    id_arquivo = os.path.splitext(os.path.basename(file_path))[0]

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    header_idx = _find_header_row(ws)
    if header_idx is None:
        raise ValueError(
            "Cabeçalho não encontrado na planilha.\n"
            f"Esperado: '{HEADER_MARKER}' na primeira coluna."
        )

    # Lê o cabeçalho para detectar posição das colunas de parcela
    header_row = list(ws.iter_rows(
        min_row=header_idx + 1, max_row=header_idx + 1, values_only=True
    ))[0] or []
    col_num_parcela, col_total_parcelas = _detect_parcel_cols(header_row)

    def _safe(row, idx):
        """Acessa índice com segurança (retorna None se fora do range)."""
        return row[idx] if idx is not None and idx < len(row) else None

    records = []
    for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
        if not row or row[0] is None:
            continue
        parsed_dt = _parse_date(row[0])
        if parsed_dt is None:
            continue

        rec = {
            'id_arquivo':            id_arquivo,
            'data_pagamento':        parsed_dt,
            'data_lancamento':       _parse_date(row[COL_DATA_LANCAMENTO]),
            'estabelecimento':       _str(row[COL_ESTABELECIMENTO], 30),
            'tipo_lancamento':       _str(row[COL_TIPO_LANCAMENTO], 100),
            'forma_pagamento':       _str(row[COL_FORMA_PAGAMENTO], 100),
            'bandeira':              _str(row[COL_BANDEIRA], 50),
            'valor_bruto':           _parse_decimal(row[COL_VALOR_BRUTO]),
            'taxa_tarifa':           _parse_decimal(row[COL_TAXA_TARIFA]),
            'valor_liquido':         _parse_decimal(row[COL_VALOR_LIQUIDO]),
            'status_pagamento':      _str(row[COL_STATUS_PAGAMENTO], 100),
            'data_venda':            _parse_date(row[COL_DATA_VENDA]),
            'hora_venda':            _str(row[COL_HORA_VENDA], 10),
            'autorizacao':           _str(row[COL_AUTORIZACAO], 30),
            'nsu_doc':               _str(row[COL_NSU_DOC], 30),
            'cod_venda':             _str(row[COL_COD_VENDA], 60),
            'numero_parcela':        _parse_int(_safe(row, col_num_parcela)),
            'total_parcelas':        _parse_int(_safe(row, col_total_parcelas)),
            'taxa_total':            _parse_decimal(_safe(row, COL_TAXA_TOTAL)),
            'valor_taxa_mdr':        _parse_decimal(_safe(row, COL_VALOR_TAXA_MDR)),
            'valor_total_transacao': _parse_decimal(_safe(row, COL_VALOR_TOTAL_TRANSACAO)),
        }
        records.append(rec)

    wb.close()
    return id_arquivo, records
